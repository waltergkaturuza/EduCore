"""
Utility functions for School Admin operations.
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_pdf_report(title, data, columns, filename='report.pdf'):
    """Generate a PDF report using ReportLab."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [columns]  # Header row
    
    for row in data:
        table_data.append([str(row.get(col, '')) for col in columns])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 0.3*inch))
    footer_style = ParagraphStyle(
        'CustomFooter',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1
    )
    footer_text = f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf


def generate_excel_report(title, data, columns, filename='report.xlsx'):
    """Generate an Excel report using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Title
    ws.merge_cells('A1:{}1'.format(get_column_letter(len(columns))))
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(col_name, '')
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(col_name)),
            max([len(str(row.get(col_name, ''))) for row in data] + [0])
        )
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Footer
    footer_row = len(data) + 4
    ws.merge_cells('A{}:{}'.format(footer_row, get_column_letter(len(columns)), footer_row))
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    footer_cell.font = Font(size=8, italic=True)
    footer_cell.alignment = Alignment(horizontal='center')
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def generate_ministry_student_register(tenant, academic_year):
    """Generate ZIMSEC/Ministry student register in required format."""
    from apps.students.models import Student
    from apps.academics.models import Class
    
    students = Student.objects.filter(
        tenant=tenant,
        is_deleted=False,
        current_class__academic_year=academic_year
    ).select_related('user', 'current_class', 'current_stream').order_by('current_class__level', 'student_id')
    
    # ZIMSEC format columns
    columns = [
        'Student ID', 'Admission Number', 'Surname', 'First Name', 'Middle Name',
        'Date of Birth', 'Gender', 'Class', 'Stream', 'Status'
    ]
    
    data = []
    for student in students:
        data.append({
            'Student ID': student.student_id,
            'Admission Number': student.admission_number or '',
            'Surname': student.user.last_name,
            'First Name': student.user.first_name,
            'Middle Name': student.user.middle_name or '',
            'Date of Birth': student.date_of_birth.strftime('%Y-%m-%d'),
            'Gender': student.gender.title(),
            'Class': student.current_class.name if student.current_class else '',
            'Stream': student.current_stream.name if student.current_stream else '',
            'Status': student.status.title(),
        })
    
    return generate_excel_report(
        title=f'Student Register - {academic_year.name}',
        data=data,
        columns=columns,
        filename=f'student_register_{academic_year.name}.xlsx'
    )


def generate_ministry_attendance_report(tenant, academic_year, term):
    """Generate Ministry attendance report in required format."""
    from apps.attendance.models import Attendance
    from apps.students.models import Student
    
    attendances = Attendance.objects.filter(
        tenant=tenant,
        academic_year=academic_year,
        date__gte=term.start_date,
        date__lte=term.end_date
    ).select_related('student', 'class_obj', 'stream').order_by('student__student_id', 'date')
    
    # Group by student
    student_attendance = {}
    for att in attendances:
        student_id = att.student.student_id
        if student_id not in student_attendance:
            student_attendance[student_id] = {
                'Student ID': student_id,
                'Student Name': att.student.user.full_name,
                'Class': att.class_obj.name if att.class_obj else '',
                'Total Days': 0,
                'Present': 0,
                'Absent': 0,
                'Late': 0,
                'Excused': 0,
                'Attendance %': 0,
            }
        
        student_attendance[student_id]['Total Days'] += 1
        if att.status == 'present':
            student_attendance[student_id]['Present'] += 1
        elif att.status == 'absent':
            student_attendance[student_id]['Absent'] += 1
        elif att.status == 'late':
            student_attendance[student_id]['Late'] += 1
        elif att.status == 'excused':
            student_attendance[student_id]['Excused'] += 1
    
    # Calculate percentages
    for student_id, data in student_attendance.items():
        total = data['Total Days']
        if total > 0:
            data['Attendance %'] = round((data['Present'] / total) * 100, 2)
    
    columns = ['Student ID', 'Student Name', 'Class', 'Total Days', 'Present', 'Absent', 'Late', 'Excused', 'Attendance %']
    data = list(student_attendance.values())
    
    return generate_excel_report(
        title=f'Attendance Report - {term.name} {academic_year.name}',
        data=data,
        columns=columns,
        filename=f'attendance_report_{term.name}_{academic_year.name}.xlsx'
    )




