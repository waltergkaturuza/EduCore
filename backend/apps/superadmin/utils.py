"""
Utility functions for superadmin app.
"""
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


def get_client_ip(request):
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def export_to_excel(queryset, model_name):
    """Export queryset to Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # Get model fields
    if queryset.exists():
        first_obj = queryset.first()
        fields = [field.name for field in first_obj._meta.fields if field.name != 'id']
        
        # Header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = field.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = getattr(obj, field, '')
                # Handle dates and foreign keys
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, '__str__'):
                    value = str(value)
                ws.cell(row=row, column=col).value = value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_invoice_pdf(invoice):
    """Generate PDF invoice."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("INVOICE", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Invoice details
    invoice_data = [
        ['Invoice Number:', invoice.invoice_number],
        ['Issue Date:', invoice.issue_date.strftime('%B %d, %Y')],
        ['Due Date:', invoice.due_date.strftime('%B %d, %Y')],
        ['Status:', invoice.status.upper()],
    ]
    
    invoice_table = Table(invoice_data, colWidths=[2*inch, 3*inch])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(invoice_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Bill to
    bill_to_style = ParagraphStyle(
        'BillTo',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#64748b'),
    )
    elements.append(Paragraph("<b>Bill To:</b>", styles['Heading2']))
    elements.append(Paragraph(invoice.tenant.name, styles['Normal']))
    elements.append(Paragraph(invoice.tenant.address, bill_to_style))
    elements.append(Paragraph(f"Email: {invoice.tenant.email}", bill_to_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Items table
    items_data = [
        ['Description', 'Amount'],
        ['Subscription Fee', f"${invoice.total:.2f}"],
    ]
    
    if invoice.tax > 0:
        items_data.append(['Tax', f"${invoice.tax:.2f}"])
    if invoice.discount > 0:
        items_data.append(['Discount', f"-${invoice.discount:.2f}"])
    
    items_data.append(['<b>TOTAL</b>', f"<b>${invoice.total:.2f}</b>"])
    
    items_table = Table(items_data, colWidths=[4*inch, 2*inch])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f8fafc')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(items_table)
    
    # Notes
    if invoice.notes:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph("<b>Notes:</b>", styles['Heading3']))
        elements.append(Paragraph(invoice.notes, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()




