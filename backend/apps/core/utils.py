"""
Utility functions for EduCore.
"""
import csv
import openpyxl
from io import StringIO, BytesIO
from typing import List, Dict


def generate_csv(data: List[Dict], headers: List[str]) -> str:
    """Generate CSV content from data."""
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def generate_excel(data: List[Dict], headers: List[str], sheet_name: str = 'Sheet1') -> BytesIO:
    """Generate Excel file from data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_phone_number(phone: str) -> str:
    """Parse and normalize phone number for Zimbabwe."""
    # Remove all non-digit characters
    phone = ''.join(filter(str.isdigit, phone))
    
    # Handle Zimbabwe phone numbers
    if phone.startswith('263'):
        return phone
    elif phone.startswith('0'):
        return '263' + phone[1:]
    elif len(phone) == 9:
        return '263' + phone
    else:
        return phone




